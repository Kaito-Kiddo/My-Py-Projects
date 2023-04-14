"""     IP addr details using AbuseIPdb.com

This script allows the user to 
open a website https://www.abuseipdb.com/check/ and check the abuse of a given IP addr

save the result like source IP, ISP, Country, Abuse Confidence % and no of times IP was reported to a Dictionary.

Create an Excel file and add Dict keys and values to excel sheet where Keys are columns and values are rows

"""
#%%
import webbrowser
import requests,bs4
import openpyxl

# s=input() '185.20.6.61' '49.36.132.29' 
s= '1.1.1.1'                                            #   Input IP
ip_dict= {'Source IP' : s}                                  #   Create Dict with first key values of Source IP addr
webbrowser.open("https://www.abuseipdb.com/check/"+s)     #   Open web browser and search for ip

res=requests.get("https://www.abuseipdb.com/check/"+s)      #   source code of web page
res.raise_for_status()
soup=bs4.BeautifulSoup(res.text,features="lxml")     

# find Confidence and no of times IP reported from ipAbuse soup 
ipAbuse=soup.select('div b')        #   Length of ipAbuse List = 17

if len(ipAbuse) == 17:
    timesReported = ipAbuse[6].get_text() + ' Times'
    confidencePercent = ipAbuse[7].get_text()
else:
    timesReported = 'Not Found'
    confidencePercent = 'Not Found'
ip_dict.update(zip(['IP reported','Confidence'],[timesReported,confidencePercent]))      # Update Dict 

# print('Length : ',len(ipAbuse))
# print('IP reported : ',timesReported,' \nConfidence : ',confidencePercent)

# ip_table contains soup of table tag class table
ip_table = soup.find("table", attrs={"class": "table"})     #   get the ip table from website

col=[]
val=[]
i=ip_table.find_all("tr")
print(len(i))
for j in i:
    if j.th.get_text() == 'Hostname(s)':
        continue
    col.append(j.th.get_text('',strip=True))
    val.append(j.td.get_text('',strip=True))
    ip_dict.update((zip(col,val)))                 # Update Dict


for key,values in ip_dict.items():
    print(key,values,sep=" : ")

        #   Input Dict values to EXCEL sheet
wb = openpyxl.load_workbook('testipinfo.xlsx')
ws = wb['IP Info'] 
print(ws.title)
for i, j in zip(ip_dict.keys(), range(1,9)):
    print(ip_dict[i])
    ws.cell(row=1,column=j).value = i
    ws.cell(row=3,column=j).value = ip_dict[i]


wb.save('testipinfo.xlsx')